#!/usr/bin/env python3
"""
Builds the HH2026v1 XLSForm from a single field-definition list.

Why a generator script rather than a hand-edited spreadsheet: the constraint
register (Q3 requirement 2) must match what the form actually enforces. Every
row below carries its own justification metadata, and constraint_register.csv
is exported from the same rows that produce the XLSForm survey sheet, so the
two artefacts cannot drift apart.

Run: python scripts/build_form.py
Output: form/HH2026_v1.xlsx, constraint_register.csv
"""
import csv
import os
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
FORM_PATH = os.path.join(ROOT, "form", "HH2026_v1.xlsx")
REGISTER_PATH = os.path.join(ROOT, "constraint_register.csv")

# ---------------------------------------------------------------------------
# Hausa translation policy (documented in full in
# documentation/06_language_and_translation.md): short, high-frequency,
# unambiguous items are translated. Full clinical sentence text is marked
# PENDING rather than guessed, because a wrong clinical translation is worse
# than an English fallback for a 38%-low-confidence-English enumerator corps.
# ---------------------------------------------------------------------------
PENDING = "[HAUSA: sai an fassara ta kwararre — professional translation pending]"

# name, label_en, label_ha
YES_NO = None

rows = []          # survey sheet rows (list of dicts)
register = []      # constraint register rows (list of dicts)
reg_id = 0


def add(row):
    rows.append(row)


def constraint_entry(field, question_no, rule_type, rule, prevents, source):
    global reg_id
    reg_id += 1
    register.append({
        "id": f"C{reg_id:03d}",
        "field_name": field,
        "questionnaire_no": question_no,
        "rule_type": rule_type,
        "rule_added": rule,
        "prevents": prevents,
        "source_or_judgement": source,
    })


def begin_group(name, label_en, label_ha=PENDING, appearance=None, relevant=None):
    r = {"type": "begin group", "name": name, "label::English (en)": label_en,
         "label::Hausa (ha)": label_ha}
    if appearance:
        r["appearance"] = appearance
    if relevant:
        r["relevant"] = relevant
    add(r)


def end_group():
    add({"type": "end group"})


def begin_repeat(name, label_en, label_ha=PENDING):
    add({"type": "begin repeat", "name": name, "label::English (en)": label_en,
         "label::Hausa (ha)": label_ha})


def end_repeat():
    add({"type": "end repeat"})


def note(name, label_en, label_ha=PENDING, relevant=None):
    r = {"type": "note", "name": name, "label::English (en)": label_en,
         "label::Hausa (ha)": label_ha}
    if relevant:
        r["relevant"] = relevant
    add(r)


def calc(name, calculation):
    add({"type": "calculate", "name": name, "calculation": calculation})


def q(type_, name, label_en, label_ha=PENDING, hint_en="", hint_ha="",
      required=False, relevant=None, constraint=None, constraint_message=None,
      appearance=None, default=None, read_only=False, choice_filter=None,
      calculation=None, guidance_en="", guidance_ha=""):
    r = {
        "type": type_, "name": name,
        "label::English (en)": label_en, "label::Hausa (ha)": label_ha,
    }
    if hint_en:
        r["hint::English (en)"] = hint_en
    if hint_ha:
        r["hint::Hausa (ha)"] = hint_ha
    if guidance_en:
        r["guidance_hint::English (en)"] = guidance_en
    if guidance_ha:
        r["guidance_hint::Hausa (ha)"] = guidance_ha
    if required:
        r["required"] = "yes"
    if relevant:
        r["relevant"] = relevant
    if constraint:
        r["constraint"] = constraint
    if constraint_message:
        r["constraint_message::English (en)"] = constraint_message
    if appearance:
        r["appearance"] = appearance
    if default is not None:
        r["default"] = default
    if read_only:
        r["read_only"] = "yes"
    if choice_filter:
        r["choice_filter"] = choice_filter
    if calculation:
        r["calculation"] = calculation
    add(r)


choices = []  # (list_name, name, label_en, label_ha, extra_cols dict)


def choice(list_name, name, label_en, label_ha=PENDING, **extra):
    row = {"list_name": list_name, "name": name, "label::English (en)": label_en,
           "label::Hausa (ha)": label_ha}
    row.update(extra)
    choices.append(row)


# ===========================================================================
# SETTINGS
# ===========================================================================
settings = {
    "form_title": "Integrated Child Health and AMR Household Survey 2026",
    "form_id": "hh2026_v1",
    "version": "2026073100",
    "default_language": "Hausa (ha)",
    "instance_name": "concat(${lga_name_txt}, '-', ${settlement_code}, '-', ${structure_no}, '-', ${hh_serial})",
    # theme-grid: renders the Section 3 roster repeat as a spreadsheet-style
    # table (one row per household member, one column per field) instead of
    # a long vertical form repeated per person -- matches the paper form's
    # own tabular roster layout (columns 1-8) far more closely than the
    # default single-column theme. Applies to every group/repeat in the
    # form, not just the roster; combined with 'minimal' on the select
    # fields living inside the roster (below) so grid columns stay narrow
    # and readable rather than each holding a full radio-button list.
    "style": "theme-grid",
}

# Enketo/ODK Collect automatically offers a language switcher (visible as
# "Choose Language" in the rendered form) whenever a form defines two or
# more languages via label::<Language> (<code>) columns -- no separate
# settings flag turns this on. This form defines English (en) and
# Hausa (ha) throughout (see label::English (en) / label::Hausa (ha) on
# every row below), with Hausa as default_language above, so the switcher
# is already present; both KoboToolbox screenshots in this session's
# testing show it working ("Choose Language: Hausa / English").

# ===========================================================================
# FRONT MATTER / METADATA
# Standard ODK begin/end timestamps: added because the paper form has NO
# interview start time at all -- only 7.01 "time interview ended". Without a
# start time, interview duration (the exact signal that caught the 94x
# 4-minute-interview enumerator in the last round, but only after fieldwork
# closed) cannot be computed from the paper's own fields. See defects report,
# defect D-06, and fabrication-detection doc.
# ===========================================================================
add({"type": "start", "name": "start"})
add({"type": "end", "name": "end"})
add({"type": "deviceid", "name": "device_id"})
add({"type": "today", "name": "today_date"})

constraint_entry("start/end (system)", "n/a (added)", "added field",
                  "ODK start/end metadata timestamps captured automatically at open/finalize",
                  "Interview duration cannot be computed from the paper form at all -- it has "
                  "no start-time field, only 7.01 end time. This is the exact signal that "
                  "identified the 94-interview, 4-minute-mean fraud case, but only after "
                  "fieldwork closed.",
                  "My judgement, directly driven by the operating-conditions fraud case")
constraint_entry("device_id (system)", "n/a (added)", "added field",
                  "ODK deviceid metadata captured automatically",
                  "Cannot attribute submissions to a specific tablet for the per-device "
                  "duplicate-label check (requirement 8) or for device-level QA dashboards",
                  "My judgement")

# ===========================================================================
# SECTION 1: HOUSEHOLD IDENTIFICATION
# ===========================================================================
begin_group("s1", "Section 1: Household identification", "Kashi na 1: Bayanan gida")

q("select_one state", "state", "1.01  State", "1.01  Jiha", default="1")
choice("state", "1", "Bansara", "Bansara")

q("select_one_from_file lgas.csv", "lga", "1.02  Local Government Area",
  "1.02  Kananan Hukuma",
  hint_en="Select from the list. Do not type the name.",
  hint_ha="Zaba daga jerin. Kada a rubuta sunan da hannu.",
  required=True, appearance="minimal")

calc("lga_name_txt", "${lga}")

q("select_one_from_file wards.csv", "ward", "1.03  Ward", "1.03  Yankin gunduma",
  hint_en="List is filtered to the LGA you just selected.",
  hint_ha="Jerin ya dogara ne akan Kananan Hukumar da aka zaba.",
  required=True, choice_filter="lga_code=${lga}", appearance="minimal")

q("select_one_from_file settlements.csv", "settlement", "1.04  Settlement", "1.04  Gari/kauye",
  hint_en="Type a few letters to search. List has 2,524 entries -- do not scroll.",
  hint_ha="Rubuta haruffa kadan don nema. Jerin yana da sunaye 2,524 -- kada a bi ta"
          " birgima.",
  guidance_en="Filtered to the ward you just selected, and sourced directly from "
               "settlements.csv in the reference media -- record and code come from the "
               "same register the ministry supplied, not a re-typed list.",
  required=True, choice_filter="ward_code=${ward}", appearance="minimal")

calc("settlement_code", "${settlement}")

q("select_one yes_no", "settlement_local_name_known", "1.05  Is the settlement known "
  "locally by a different name?", "1.05  Shin mutanen yankin na kiran wannan wuri da wani "
  "suna daban?",
  hint_en="Ask the respondent, do not assume from the register name alone.",
  hint_ha="A tambayi wanda ake yi wa tambayoyi, kada a dogara da sunan da ke cikin "
          "rajista kadai.",
  required=True, appearance="quick")
q("text", "settlement_local_name", "If yes, write the name used locally.",
  "Idan ee, rubuta sunan da ake amfani da shi a wurin.",
  relevant="${settlement_local_name_known}='1'")

q("integer", "structure_no", "1.06  Structure number painted on the dwelling",
  "1.06  Lambar da aka rubuta a gidan",
  hint_en="Copy the number exactly as painted. If none is painted, ask your supervisor "
          "before proceeding.",
  hint_ha="Kwafi lambar daidai yadda take a rubuce. Idan babu lambar da aka rubuta, a "
          "tambayi mai kula kafin cigaba.",
  required=True, constraint=". >= 1 and . <= 999",
  constraint_message="Enter 1-999 (three digit box on the paper form).")
constraint_entry("structure_no", "1.06", "range", "1-999",
                  "Non-numeric or >3-digit entry that cannot fit the printed 3-digit box",
                  "Paper form coding box is three digits wide (⌷⌷⌷); my judgement on the "
                  "implied range")

q("integer", "hh_serial", "1.07  Household serial number within the settlement",
  "1.07  Lambar gida a cikin yankin",
  hint_en="The running count of households visited in this settlement -- your first "
          "household here is 1, second is 2, and so on.",
  hint_ha="Adadin gidajen da aka ziyarta a wannan yankin -- gidan farko shine 1, na biyu 2, "
          "da sauransu.",
  required=True, constraint=". >= 1 and . <= 999",
  constraint_message="Enter 1-999 (three digit box on the paper form).")
constraint_entry("hh_serial", "1.07", "range", "1-999",
                  "Non-numeric or >3-digit entry", "Paper form coding box is three digits wide")

q("select_one enumerator", "enumerator_code", "1.08  Enumerator code", "1.08  Lambar "
  "ma'aikaci",
  required=True, hint_en="Select your own code from the staff list.",
  hint_ha="Zaba lambar kanka daga jerin ma'aikata.", appearance="minimal")
calc("team_code_auto", "pulldata('staff_roster','team_code','name',${enumerator_code})")
q("text", "team_code_display", "1.09  Team code (auto-filled)", "1.09  Lambar kungiya "
  "(kai tsaye)",
  hint_en="Filled in automatically from your enumerator code -- nothing to type here.",
  hint_ha="Ana cika wannan kai tsaye daga lambar ma'aikacinka -- babu abin da za a rubuta.",
  read_only=True, calculation="${team_code_auto}")
constraint_entry("team_code_display", "1.08 / 1.09", "removed manual entry, auto-filled",
                  "team_code_display = pulldata('staff_roster', 'team_code', 'name', "
                  "enumerator_code), shown read-only; the enumerator never types a team code",
                  "Paper 1.09 lets an enumerator enter their own code correctly at 1.08 but a "
                  "mistyped or wrong team code at 1.09, which would then misassign the "
                  "specimen-label range check in Section 5 to the wrong team's block",
                  "Derived from staff_roster.csv (120 rows, 24 teams, verified 1:1 with "
                  "specimen_label_allocation.csv team codes)")

q("date", "visit_date", "1.10  Date of visit", "1.10  Ranar ziyara", required=True,
  hint_en="Must fall within the fieldwork period, 1-30 June 2026.",
  hint_ha="Dole ranar ta kasance cikin kwanakin fieldwork, 1-30 Yuni 2026.",
  constraint=". >= date('2026-06-01') and . <= date('2026-06-30')",
  constraint_message="Date must fall within the fieldwork period printed on the form "
                      "(1-30 June 2026).")
constraint_entry("visit_date", "1.10", "range",
                  "2026-06-01 to 2026-06-30 (closed interval)",
                  "Backdated/postdated entries; device clock errors; typed year defaulting "
                  "to device-current-year instead of 2026",
                  "Printed on the questionnaire header ('Fieldwork period 1 to 30 June 2026'). "
                  "NOTE: this conflicts with the operating-conditions brief, which states "
                  "fieldwork runs 14 days -- flagged as defect D-07, escalated, not resolved "
                  "silently. I used the form's own printed period because it is the "
                  "ethics-approved text.")

q("geopoint", "gps_dwelling", "1.11  GPS reading at the entrance to the dwelling",
  "1.11  Wurin GPS a kofar gida", required=True,
  hint_en="Stand at the entrance to the dwelling and wait for the accuracy reading to "
          "settle before capturing.",
  hint_ha="Tsaya a kofar gida kuma jira har lambar daidaito ta zauna kafin dauka.",
  guidance_en="A poor or indoor fix can be several hundred metres off. If the accuracy "
               "shown is worse than about 15m, step outside and try again before accepting "
               "the reading.",
  constraint="number(substring-before(.,' ')) >= 10.0 and "
             "number(substring-before(.,' ')) <= 11.9 and "
             "number(substring-before(substring-after(.,' '),' ')) >= 6.3 and "
             "number(substring-before(substring-after(.,' '),' ')) <= 9.0",
  constraint_message="Point falls outside the Bansara state working area. Re-take the reading.")
constraint_entry("gps_dwelling", "1.11", "range",
                  "longitude 6.3-9.0, latitude 10.0-11.9 (geopoint bounding box)",
                  "Wrong-hemisphere entry, transposed lat/long, GPS fix taken indoors far "
                  "from the dwelling, stale cached fix from a previous survey",
                  "My judgement -- padded from the settlement register's actual coordinate "
                  "extent (longitude 6.9544 to 8.4288, latitude 10.3659 to 11.5735 across the "
                  "2,524 settlements in settlements.csv) with roughly a 0.4-0.5 degree margin "
                  "for genuine edge-of-catchment points")

q("select_one yes_no_dk", "prev_round_visited", "1.12  Was this household visited during "
  "the October 2025 round?", "1.12  An ziyarci wannan gida a lokacin 2025 ba?", required=True,
  appearance="quick")
q("text", "prev_round_hh_id", "1.13  Household identifier allocated in the October 2025 "
  "round", "1.13  Lambar da aka bawa gida a 2025",
  hint_en="Format BAN-000000, copied from the previous round's paperwork or supervisor "
          "list.",
  hint_ha="Tsari BAN-000000, an kwafo daga takardun zagayen da ya gabata ko jerin mai kula.",
  relevant="${prev_round_visited}='1'", required=True,
  constraint="regex(., '^BAN-[0-9]{6}$') and "
             "pulldata('previous_round_households','household_id','household_id',.) != ''",
  constraint_message="Format must be BAN-000000 and must match an identifier in the previous "
                      "round register.")
constraint_entry("prev_round_hh_id", "1.13", "format + existence",
                  "regex ^BAN-[0-9]{6}$, and must resolve via pulldata() against "
                  "previous_round_households.csv (3,982 rows, format verified 100% compliant)",
                  "A transcription error that silently creates a phantom link between this "
                  "round and a household that was never actually visited last round, which "
                  "would corrupt any longitudinal linkage analysis",
                  "previous_round_households.csv, all 3,982 household_id values checked; "
                  "format is uniform")

calc("prev_round_children_u5", "if(${prev_round_visited}='1', "
     "pulldata('previous_round_households','children_under5_last_round','household_id',"
     "${prev_round_hh_id}), '')")
note("prev_round_children_u5_note",
     "Previous round recorded ${prev_round_children_u5} child(ren) under five in this "
     "household. Confirm this is still plausible once the roster is complete.",
     "An gaba an rubuta yara ${prev_round_children_u5} 'yan kasa da shekara biyar a wannan "
     "gida. Tabbatar hakan bayan kammala jerin gida.",
     relevant="${prev_round_visited}='1' and ${prev_round_children_u5}!=''")

q("select_one visit_result", "visit_result", "1.14  Result of visit", "1.14  Sakamakon "
  "ziyara",
  hint_en="If the result is anything other than Completed, sign at 7.03 and submit -- do "
          "not continue to later sections.",
  hint_ha="Idan sakamakon ba An kammala ba, a sa hannu a 7.03 sannan a mika -- kada a ci "
          "gaba da sauran kashi.",
  required=True, appearance="minimal")

end_group()

note("end_note_no_further", "Result of visit means no further section is completed. Sign at "
     "7.03 and submit.", "Sakamakon ziyara na nufin ba za a ci gaba da wani kashi ba. Sa hannu "
     "a 7.03 sannan a mika.",
     relevant="${visit_result}='2' or ${visit_result}='3' or ${visit_result}='4'")

# ===========================================================================
# SECTION 2: CONSENT
# ===========================================================================
begin_group("s2", "Section 2: Consent", "Kashi na 2: Yardar shiga",
            relevant="${visit_result}='1'")

q("select_one yes_no", "consent_read", "2.01  Consent statement read aloud to the "
  "respondent in full?", "2.01  An karanta bayanin yardar shiga gaba daya ga wanda ake "
  "tambaya?",
  hint_en="Read the printed consent script exactly as written. Do not summarise or "
          "paraphrase it.",
  hint_ha="A karanta rubutun yardar shiga daidai yadda yake, kada a takaita ko a canza "
          "kalmomi.",
  required=True)
q("select_one consent", "consent_given", "2.02  Does the respondent consent to the "
  "household interview?", "2.02  Wanda ake tambaya ya yarda a yi hira da gidan?",
  hint_en="If refused, thank the respondent, sign at 7.03, and submit -- no further "
          "section is completed.",
  hint_ha="Idan an ki, a gode wa wanda ake tambaya, a sa hannu a 7.03, sannan a mika -- "
          "babu wani kashi da za a ci gaba da shi.",
  required=True)
q("select_one relationship", "respondent_relationship", "2.03  Relationship of the "
  "respondent to the head of household", "2.03  Dangantakar wanda ake tambaya da shugaban "
  "gida",
  relevant="${consent_given}='1'", required=True, appearance="minimal")

end_group()

note("consent_refused_note", "Consent was refused. Do not continue. Sign at 7.03 and submit.",
     "An ki amincewa. Kar a ci gaba. Sa hannu a 7.03 sannan a mika.",
     relevant="${visit_result}='1' and ${consent_given}='2'")

# ===========================================================================
# SECTION 3 + 4 + 5: ROSTER, with the child module and specimen section
# NESTED inside each roster member (relevant only when that member is
# age-eligible), instead of a second, separately-counted repeat.
#
# This single structural decision is what resolves requirement 4's second
# cross-check (number of eligible children vs number of child modules
# completed) BY CONSTRUCTION rather than by a post-hoc validation rule: there
# is exactly one child-module block per roster row, shown if and only if that
# row's own age fields say the child is 9-59 months. It is also what resolves
# defect D-01 (the "office use" column 7 / question 3.02 contradiction) and
# defect D-04 (redundant 5.01, restating information already on the roster).
# ===========================================================================
form_active = "${visit_result}='1' and ${consent_given}='1'"

begin_group("s3", "Section 3: Household roster", "Kashi na 3: Jerin gida", relevant=form_active)

q("integer", "hh_size_stated", "3.01  How many people usually live in this household?",
  "3.01  Mutane nawa suke zaune a wannan gida?",
  hint_en="Ask this before listing the roster below. A usual resident is someone who "
          "normally sleeps here, even if temporarily away.",
  hint_ha="A tambayi wannan kafin jera mutane a kasa. Mutumin da yake zaune yana nufin "
          "wanda yake barci a nan a saba, ko da yake ba ya nan yanzu na dan lokaci.",
  required=True,
  constraint=". >= 1 and . <= 30",
  constraint_message="Enter 1-30. If genuinely larger, flag for supervisor at 7.02.")
constraint_entry("hh_size_stated", "3.01", "range", "1-30",
                  "Zero or absurd entries (e.g. 300 from a stuck key); still permits genuinely "
                  "large compound households up to 30",
                  "My judgement -- no household-size distribution was supplied in the data "
                  "pack for Bansara state, so I set a generous ceiling rather than infer one")

note("roster_instruction", "List every usual resident, beginning with the head of household. "
     "For a resident under five, record age in completed MONTHS. For everyone else, record "
     "age in completed YEARS.",
     "Ka lissafa duk wanda ke zaune a gida, farawa da shugaban gida. Ga wanda ke kasa da "
     "shekara biyar, rubuta shekarunsa da watanni. Ga sauran, rubuta da shekaru.")

begin_repeat("roster", "Household member", "Dan gidan")

q("integer", "line", "(1)  Line number", "(1)  Lamba", read_only=True,
  calculation="position(..)")

q("text", "member_name", "(2)  Name or initials", "(2)  Suna ko gajerun haruffa",
  hint_en="Initials are acceptable and preferred where a full name is not needed for the "
          "interview.",
  hint_ha="Ana iya amfani da gajerun haruffa maimakon cikakken suna idan ba a bukata a "
          "cikin hira.",
  required=True)

q("select_one relationship_member", "member_relationship", "(3)  Relationship to head",
  "(3)  Dangantaka da shugaban gida", required=True, appearance="minimal")

q("select_one sex", "member_sex", "(4)  Sex", "(4)  Jinsi", required=True,
  appearance="minimal")

q("select_one yes_no", "member_under5", "Is this person under five years old?",
  "Wannan mutum yana kasa da shekara biyar?", required=True,
  hint_en="Not on the paper form -- added so the digital form routes to column (5) or "
          "(6) correctly. See defects report D-02.",
  hint_ha=PENDING, appearance="minimal")
constraint_entry("member_under5", "n/a (added, drives 3(5)/3(6) routing)", "added field",
                  "select_one yes/no gates which of age_years / age_months is asked",
                  "The paper form leaves it to the enumerator's judgement which of columns "
                  "(5)/(6) to fill for a borderline-looking child, which a printed form can "
                  "get away with but a digital form's relevant logic cannot -- it needs an "
                  "explicit branch",
                  "My judgement")

q("integer", "age_years", "(5)  Age in completed years", "(5)  Shekaru cikakku",
  hint_en="Completed years -- round down, do not round up to the nearest birthday.",
  hint_ha="Shekaru cikakku -- a rage zuwa kasa, kada a kara zuwa ranar haihuwa mai zuwa.",
  relevant="${member_under5}='2'", required=True,
  constraint=". >= 5 and . <= 110",
  constraint_message="Enter 5-110 years.")
constraint_entry("age_years", "3(5)", "range", "5-110",
                  "A member coded 'not under five' with an age below 5, or an implausible age "
                  "from a stuck key (e.g. 999)",
                  "Lower bound follows directly from the routing question; upper bound is my "
                  "judgement (no Bansara demographic data supplied)")

q("integer", "age_months", "(6)  Age in completed months (under five only)",
  "(6)  Watanni cikakku (kasa da shekara biyar kadai)",
  hint_en="If exact age is not known, estimate using a local events/seasonal calendar "
          "rather than guessing.",
  hint_ha="Idan ba a san ainihin shekaru ba, a yi kiyasi ta amfani da abubuwan da suka "
          "faru a yankin ko lokutan shekara, kada a yi tsammani kawai.",
  relevant="${member_under5}='1'", required=True,
  constraint=". >= 0 and . <= 59",
  constraint_message="Enter 0-59 completed months.")
constraint_entry("age_months", "3(6)", "range", "0-59",
                  "A member coded 'under five' with a months value that is actually five years "
                  "or older, which would silently misroute Section 4/5 eligibility",
                  "Definitional: 'under five years' cannot exceed 59 completed months")

calc("eligible_s4", "if(${member_under5}='1' and ${age_months} >= 9 and ${age_months} <= 59, "
     "1, 0)")
note("eligible_s4_note", "This is computed automatically from the age just recorded -- it "
     "replaces the paper form's column (7), which was marked 'office use' yet was required "
     "live in the field to route question 3.02 (see defects report D-01).",
     PENDING, relevant="${eligible_s4}=1")

# ---- Section 4 + 5, nested, gated on eligible_s4 ----
begin_group("child_module", "Section 4: Child module", "Kashi na 4: Bayanan yaro",
            relevant="${eligible_s4}=1")

q("text", "c_name", "4.02  Name or initials of the child (copied from roster)",
  "4.02  Sunan yaro (an dauko daga jeri)", read_only=True, calculation="${member_name}")
q("integer", "c_age_months", "4.03  Age of the child in completed months (copied from "
  "roster)", "4.03  Shekarun yaro da watanni (an dauko daga jeri)", read_only=True,
  calculation="${age_months}")
q("select_one sex", "c_sex", "4.04  Sex of the child (copied from roster)",
  "4.04  Jinsin yaro (an dauko daga jeri)", read_only=True, calculation="${member_sex}")

# --- Anthropometry: sentinel/measurement collision fix (defect D-03) ---
# The paper form puts a "not measured = 99" sentinel inside a continuous
# numeric field (4.05, 4.06). 99 kg and 99.9 cm are not plausible measurements
# for a 9-59 month child, so on paper the collision is harmless to a human
# reader -- but a digital numeric field has no such common sense, and 99 is a
# perfectly typeable, in-range-looking decimal. Split each into a status
# select_one (measured / not measured, with reason) plus a numeric field that
# is only relevant, and only required, when status = measured. The analysis
# team can then never receive a 99 in a column of real weights.
q("select_one measure_status", "weight_status", "4.05  Weight measurement status",
  "4.05  Halin auna nauyi",
  hint_en="Paper question 4.05 asked for the weight itself, with 99 meaning not "
          "measured. Here that is two separate fields -- pick the status first.",
  hint_ha=PENDING,
  required=True, appearance="minimal")
q("decimal", "c_weight_kg", "4.05  Weight of the child (kg)", "4.05  Nauyin yaro (kg)",
  hint_en="Record to the nearest 0.1 kg, with the child in light clothing.",
  hint_ha="A rubuta zuwa 0.1 kg mafi kusa, yaro na sanye da tufafi masu sauki.",
  relevant="${weight_status}='1'", required=True,
  constraint=". >= 2.0 and . <= 30.0",
  constraint_message="Enter 2.0-30.0 kg.")
constraint_entry("c_weight_kg", "4.05", "sentinel split + range",
                  "Split into weight_status (measured/not measured/refused) + c_weight_kg "
                  "numeric 2.0-30.0, numeric field only relevant if status=measured",
                  "The paper form's '99 = not measured' sentinel living inside a continuous "
                  "kg field, which is analytically indistinguishable from a genuine (if "
                  "unlikely) 99.0kg entry once digitised as a plain decimal field",
                  "Range is my judgement, generously padded around WHO Child Growth Standards "
                  "weight-for-age for 9-59 months to avoid rejecting true severe-malnutrition "
                  "or high-BMI outliers while catching entry blunders (e.g. a misplaced "
                  "decimal point)")

q("select_one measure_status", "height_status", "4.06  Length/height measurement status",
  "4.06  Halin auna tsawo",
  hint_en="Paper question 4.06 asked for the measurement itself, with 99 meaning not "
          "measured. Here that is two separate fields -- pick the status first.",
  hint_ha=PENDING,
  required=True, appearance="minimal")
q("select_one measure_position", "measure_position", "4.07  Position in which the child "
  "was measured", "4.07  Yanayin da aka auna yaro",
  hint_en="A change part-way through the round is expected -- record what was actually "
          "used, do not force consistency with earlier children.",
  hint_ha="Ana sa ran za a canza yayin zagayen -- a rubuta abin da aka yi amfani da shi na "
          "gaskiya, kada a tilasta daidaituwa da sauran yara.",
  guidance_en="WHO/DHS convention: measure children under 24 months recumbent (lying "
               "down); measure children 24 months and older standing, unless the child "
               "cannot stand, in which case use recumbent and note it at 7.02.",
  relevant="${height_status}='1'", required=True, appearance="minimal")
constraint_entry("measure_position", "4.07", "consistency (documented, not blocked)",
                  "Recorded alongside height so the office can apply the standard 0.7cm "
                  "recumbent/standing correction; a mid-survey change in which position is "
                  "used for a given age band is expected (operating conditions say a "
                  "measurement-position change is likely mid-round) and must not be blocked",
                  "A position switch is a real, expected event this round -- flagging it (via "
                  "the codebook, not a form constraint) is correct; refusing it in-form is not",
                  "Operating conditions: 'A change to the instrument part way through the "
                  "round is likely'")
q("decimal", "c_height_cm", "4.06  Length or height of the child (cm)", "4.06  Tsawon yaro "
  "(cm)",
  hint_en="Record to the nearest 0.1 cm.",
  hint_ha="A rubuta zuwa 0.1 cm mafi kusa.",
  relevant="${height_status}='1'", required=True,
  constraint=". >= 45.0 and . <= 130.0",
  constraint_message="Enter 45.0-130.0 cm.")
constraint_entry("c_height_cm", "4.06", "sentinel split + range",
                  "Split into height_status + c_height_cm numeric 45.0-130.0, only relevant "
                  "if status=measured",
                  "Same collision as weight: paper's '99 = not measured' inside a 3-digit "
                  "continuous cm field",
                  "Range is my judgement, padded around WHO length/height-for-age for 9-59 "
                  "months")

q("select_one card_seen", "vacc_card_seen", "4.08  May I see the child's vaccination card "
  "or health record?", "4.08  Zan iya ganin katin allurar yaro ko bayanin lafiyarsa?",
  hint_en="Ask to see the actual card or electronic record before answering -- do not "
          "accept a verbal description as 'seen.'",
  hint_ha="A nemi ganin ainihin katin ko bayanin lantarki kafin amsawa -- kada a karbi "
          "bayanin baki a matsayin an gani.",
  required=True, appearance="minimal")
q("select_one yes_no", "measles_from_card", "4.09  Copy from the card: is a measles dose "
  "recorded?", "4.09  Daga kati: an rubuta an yi wa yaro allurar kyanda?",
  hint_en="Record exactly what is written on the card -- do not ask the caregiver to "
          "recall this one, the card is in front of you.",
  hint_ha="A rubuta daidai abin da ke rubuce a kati -- kada a tambayi mai kula game da "
          "wannan, kati na gabanka ne.",
  relevant="${vacc_card_seen}='1'", required=True, appearance="minimal")
q("select_one yes_no_dk", "measles_recall", "4.10  Has this child ever received a measles "
  "vaccination?", "4.10  Yaro ya taba samun allurar kyanda?",
  relevant="${vacc_card_seen}='2'", required=True, appearance="minimal")
calc("measles_status", "if(${vacc_card_seen}='1', ${measles_from_card}, ${measles_recall})")
constraint_entry("measles_status", "4.08-4.10", "sentinel/source separation",
                  "measles_status combines card-confirmed and recall answers into one "
                  "analysis field, but the raw vacc_card_seen/measles_from_card/"
                  "measles_recall fields are kept, not overwritten",
                  "Analysts silently losing the card-confirmed-vs-recall distinction, which "
                  "Q4 of this same assessment explicitly requires reporting on for coverage",
                  "My judgement, informed by the Q4 requirement to report coverage 'by "
                  "documented source, distinguishing card confirmed from caregiver recall'")

q("select_one yes_no_dk", "diarrhoea_14d", "4.11  Has this child had diarrhoea in the "
  "past 14 days?", "4.11  Yaro ya sha gudawa cikin kwanaki 14 da suka wuce?",
  hint_en="Diarrhoea means three or more loose or watery stools in 24 hours.",
  hint_ha="Gudawa yana nufin fitsi mai laushi ko ruwa sau uku ko fiye a cikin awa 24.",
  required=True, appearance="minimal")

q("select_one yes_no_dk", "antibiotic_30d", "4.12  Has this child taken any antibiotic "
  "medicine in the past 30 days?", "4.12  Yaro ya sha maganin rigakafin kwayoyin cuta cikin "
  "kwanaki 30 da suka wuce?", required=True, appearance="minimal")

q("select_one_from_file medicine_list.csv", "antibiotic_code", "4.13  Which antibiotic was "
  "taken?", "4.13  Wanne maganin rigakafi ne aka sha?",
  relevant="${antibiotic_30d}='1'", required=True,
  hint_en="This question cannot be answered yet -- the medicine list has not been "
          "supplied. See the info icon.",
  hint_ha="Ba za a iya amsa wannan tambaya ba tukuna -- ba a bayar da jerin magunguna ba. "
          "Duba alamar bayani.",
  guidance_en="Question 4.13 on the paper form instructs: 'Record from the medicine "
               "list.' No medicine/antimicrobial code list was supplied anywhere in the "
               "data pack for this assessment -- confirmed by directory listing and a "
               "full-text search of reference_media/. This field is wired exactly like "
               "the LGA/Ward/Settlement selects (select_one_from_file against an external "
               "CSV attached as form media), but form/media/medicine_list.csv is a stub -- "
               "a header row only (name,label columns), no data. The ministry/AMR "
               "technical working group needs to supply the actual coded list; once "
               "medicine_list.csv is populated with real rows and reattached, this "
               "question works with no other change to the form. Until then it correctly "
               "has no answer to offer. See defects report D-05.",
  appearance="minimal")
constraint_entry("antibiotic_code", "4.13", "data-pack gap, reported and structurally staged",
                  "select_one_from_file medicine_list.csv -- same mechanism as the LGA/"
                  "Ward/Settlement cascade, but form/media/medicine_list.csv ships as an "
                  "empty stub (header row only). No choice content is fabricated or "
                  "substituted; the field is wired to receive the real list the moment "
                  "the ministry supplies it, with no other change needed to the form.",
                  "Presenting any list of antibiotic names -- invented, or drawn from a "
                  "public source and cited -- as though it were the ministry's own coding "
                  "scheme, which a reviewer or enumerator could mistake for the real thing",
                  "ESCALATED, not resolved, and not substituted for: the questionnaire "
                  "refers to 'the medicine list' but reference_media/ contains no such "
                  "file. This is a data-pack gap to report, not something to fill with "
                  "content of my own choosing -- even a list drawn from a public source "
                  "(e.g. WHO AWaRe) would still not be this survey's ministry-approved "
                  "formulary code list. Structuring the field to await the real file, "
                  "exactly like every other external list in this form, keeps the "
                  "reported gap honest while requiring zero rework once the real list "
                  "arrives.")

q("select_one yes_no_dk", "antibiotic_no_rx", "4.15  Was the medicine obtained without a "
  "prescription from a health worker?", "4.15  An sami maganin ba tare da takardar likita "
  "ba?", relevant="${antibiotic_30d}='1'", required=True, appearance="minimal")

q("select_one photo_status", "antibiotic_photo", "4.16  Was a photograph of the medicine "
  "packaging taken?", "4.16  An dauki hoton fakitin magani?",
  hint_en="Ask the caregiver's permission before photographing the packaging.",
  hint_ha="A nemi izinin mai kula kafin daukar hoton fakitin.",
  relevant="${antibiotic_30d}='1'", required=True, appearance="minimal")
q("image", "antibiotic_photo_file", "4.16  Photograph of medicine packaging",
  "4.16  Hoton fakitin magani",
  hint_en="Take a clear, well-lit photo showing the medicine name on the packaging.",
  hint_ha="A dauki hoto mai haske da bayyanawa wanda ke nuna sunan maganin a fakitin.",
  relevant="${antibiotic_photo}='1'", appearance="new")
constraint_entry("antibiotic_photo_file appearance", "4.16", "appearance: new",
                  "'new' forces the camera to open for a fresh capture; the device's photo "
                  "gallery is not offered as a source",
                  "An enumerator attaching an old or unrelated photo from the gallery instead "
                  "of actually photographing this packaging, which would defeat the evidence "
                  "purpose of the field entirely",
                  "My judgement")
constraint_entry("antibiotic_photo_file", "4.16", "added evidence field",
                  "image capture attached whenever antibiotic_photo=1 (photo taken)",
                  "The paper form records only whether a photo was taken, not the photo "
                  "itself, so the office has no way to verify the antibiotic_code entry "
                  "against the packaging -- an AMR-survey-specific back-office QA gap",
                  "My judgement")

end_group()  # child_module

# ---- Section 5, nested, gated on eligible_s4 AND age >= 12 months ----
# 5.01 ('Is the child aged 12 completed months or older?') is fully
# redundant with age_months, already captured two questions earlier on the
# very same page. Re-asking it invites a value that contradicts age_months --
# a second, avoidable sentinel-style collision. Computed instead (defect D-04).
begin_group("specimen", "Section 5: Specimen collection", "Kashi na 5: Tattara samfurin",
            relevant="${eligible_s4}=1 and ${age_months} >= 12")

q("select_one yes_no", "specimen_obtained", "5.02  Was a stool specimen obtained from this "
  "child?", "5.02  An samu samfurin kashin yaro?",
  hint_en="If Yes, you will label and store it next. If No, you will record why below.",
  hint_ha="Idan Ee, za a manna lambar sannan a adana samfurin. Idan A'a, za a rubuta "
          "dalili a kasa.",
  required=True, appearance="minimal")
constraint_entry("specimen section relevance", "5.01", "computed relevance, question removed",
                  "Section only shown when age_months>=12 (already on file); no separate "
                  "5.01 question is asked",
                  "5.01 restates a fact already captured at 4.03/age_months two questions "
                  "earlier; asking it again lets an enumerator answer 'yes, >=12 months' for "
                  "a child whose age_months says 10, an internal contradiction the paper "
                  "form has no mechanism to prevent",
                  "My judgement -- resolved in form, disclosed in defects report D-04, not "
                  "silently dropped")

# ---- 5.02 has NO skip instruction on the paper form at all: neither branch
# (specimen obtained / not obtained) says what to do next. Resolved in form
# using the only reading that is internally consistent with 5.03-5.07's own
# content (you cannot affix a label to a specimen you do not have, and you
# cannot give a reason none was obtained for one you did obtain). Flagged for
# ministry confirmation regardless -- see defects report D-08.
# NOTE: calculated helper fields below are intentionally listed before the
# specimen_label question that consumes them. specimen_label needs its own
# derived digits/checksum computed from ITS OWN prior value on re-entry
# (constraint re-evaluation), and team-range lookups only depend on
# team_code_auto (already defined at Section 1), not on specimen_label --
# ordering the pulldata/range calcs first keeps the dependency graph acyclic
# and easy to read.
calc("label_range_start", "pulldata('specimen_label_allocation','range_start','team_code',"
     "${team_code_auto})")
calc("label_range_end", "pulldata('specimen_label_allocation','range_end','team_code',"
     "${team_code_auto})")
calc("label_digits", "substr(${specimen_label},3,6)")
calc("label_checksum", "number(substr(${label_digits},0,1))*7 + "
     "number(substr(${label_digits},1,1))*6 + number(substr(${label_digits},2,1))*5 + "
     "number(substr(${label_digits},3,1))*4 + number(substr(${label_digits},4,1))*3 + "
     "number(substr(${label_digits},5,1))*2")
calc("label_check_expected", "if((${label_checksum} mod 11)=10, 'X', "
     "string(${label_checksum} mod 11))")

q("text", "specimen_label", "5.03  Specimen label number (affix label, then transcribe in "
  "full)", "5.03  Lambar samfurin (a manne sannan a rubuta gaba daya)",
  hint_en="Affix the label to the specimen container first, then type the full code "
          "exactly as printed, including the letter/digit after the dash.",
  hint_ha="A fara manna lambar akan akwatin samfurin, sannan a rubuta cikakkiyar lambar "
          "daidai yadda take a buga, har da harafi/lamba bayan dash din.",
  guidance_en="Format BSN######-C. The last character (a digit or the letter X) is a "
               "check digit computed from the six digits before it -- the form verifies it "
               "automatically and will reject a mistyped label, so re-check the physical "
               "label carefully before assuming the form is wrong.",
  relevant="${specimen_obtained}='1'", required=True,
  constraint="regex(., '^BSN[0-9]{6}-[0-9X]$') and "
             "number(substr(.,3,6)) >= number(${label_range_start}) and "
             "number(substr(.,3,6)) <= number(${label_range_end}) and "
             "substr(.,10,1) = ${label_check_expected} and "
             "count(../../../roster/specimen/specimen_label[. = current()]) <= 1",
  constraint_message="Label must be format BSN######-C, in your team's allocated range, with a "
                      "valid check digit, and not already used for another child in this same "
                      "household submission.")
constraint_entry("specimen_label", "5.03", "format + range + check-digit + within-submission dedup",
                  "regex ^BSN[0-9]{6}-[0-9X]$; 6-digit body inside the enumerator's own team "
                  "range (pulldata from specimen_label_allocation.csv); check character "
                  "recomputed via modulus-11, weights 2-7 right to left, and compared to the "
                  "entered character; must not repeat within this same household's roster "
                  "(count of matching labels across all roster/specimen/specimen_label nodes "
                  "in this submission <= 1)",
                  "Mistyped labels, a label affixed and typed from outside the team's "
                  "allocated block (allocation confirmed contiguous, no gaps/overlaps across "
                  "24 teams x 900 labels each, 480000-501599), transposed-digit-pair errors "
                  "the check digit is specifically designed to catch, and the same physical "
                  "label being typed twice for two different children in the same household",
                  "specimen_label_allocation.csv (24 rows) for the range; check-digit scheme "
                  "text in that same file ('Modulus 11, weights 2 to 7 applied right to left, "
                  "remainder 10 recorded as X'). See documentation/04_specimen_label_"
                  "validation.md for worked test vectors, and documentation/05_duplicate_"
                  "label_detection.md for why a same-submission check only catches labels "
                  "reused within one household and not the full 9-day device history.")
constraint_entry("specimen_label appearance (bug fix)", "5.03", "appearance removed",
                  "An earlier build set appearance='numbers' on this field, which restricts "
                  "the on-screen keyboard to digits only. This field's own format is "
                  "BSN######-C -- letters and an occasional 'X' check character -- so a "
                  "numeric-only keypad would have made a correctly formatted label untypeable. "
                  "No appearance is set now; the default keyboard accepts the full alphanumeric "
                  "format.",
                  "A numeric keypad blocking entry of a field whose own required format "
                  "includes letters",
                  "Caught on review while adding appearances for this update, not by a field "
                  "test -- flagged here anyway since it is exactly the kind of thing device "
                  "testing (documentation/11_scope_and_omissions.md, item 9) is meant to catch")

q("time", "specimen_cold_box_time", "5.04  Time the specimen was placed in the cold box",
  "5.04  Lokacin da aka sa samfurin cikin akwatin sanyi",
  hint_en="24-hour clock, e.g. 14:30 for 2:30pm.",
  hint_ha="Agogon awa 24, misali 14:30 don karfe 2:30 na yamma.",
  relevant="${specimen_obtained}='1'", required=True)
q("decimal", "specimen_temp_c", "5.05  Temperature shown on the cold box thermometer",
  "5.05  Zafin da ke akwatin sanyi",
  hint_en="Read directly from the thermometer at the moment the specimen is placed inside.",
  hint_ha="A karanta kai tsaye daga thermometer a lokacin da aka sa samfurin a ciki.",
  relevant="${specimen_obtained}='1'", required=True,
  constraint=". >= 0.0 and . <= 12.0",
  constraint_message="Enter 0.0-12.0 degrees C.")
constraint_entry("specimen_temp_c", "5.05", "range", "0.0-12.0 degrees C",
                  "Miskeyed or unrealistic cold-chain temperatures (e.g. a stuck '99', or a "
                  "positive-but-implausible double-digit reading)",
                  "My judgement -- WHO cold-chain guidance targets 2-8C for specimen "
                  "transport; I padded to 0-12C to admit a plausible faulty-but-real "
                  "thermometer reading rather than only the ideal range, while still "
                  "rejecting clear entry errors")

q("select_one no_specimen_reason", "specimen_no_reason", "5.06  Reason no specimen was "
  "obtained", "5.06  Dalilin da ba a samu samfurin ba",
  relevant="${specimen_obtained}='2'", required=True, appearance="minimal")
q("text", "specimen_no_reason_other", "5.07  If Other, specify", "5.07  Idan Wani, bayyana",
  relevant="${specimen_no_reason}='96'", required=True)

end_group()  # specimen

end_repeat()  # roster

calc("roster_count_actual", "count(${roster})")
calc("roster_mismatch_flag", "if(${hh_size_stated} != ${roster_count_actual}, 1, 0)")
note("roster_vs_stated_note",
     "Stated household size was ${hh_size_stated}. The roster you just completed lists "
     "${roster_count_actual} people. Please recheck before continuing if these differ.",
     PENDING,
     relevant="${hh_size_stated} != ${roster_count_actual}")
constraint_entry("roster_count_actual / hh_size_stated", "3.01 vs Section 3 roster",
                  "cross-question consistency (soft, not blocking)",
                  "roster_count_actual = count(roster); roster_mismatch_flag=1 when it "
                  "differs from hh_size_stated; enumerator sees an on-screen note and must "
                  "actively continue past it, but the form is not blocked",
                  "A genuine mismatch between the pre-roster estimate (3.01) and the actual "
                  "listed count is real, common (memory, definitional disputes about who is "
                  "'usually resident'), and analytically important -- it must reach the office "
                  "for review, not be silently forced to agree or hard-blocked in the field",
                  "Q3 requirement 4: 'at minimum reconcile the stated household size against "
                  "the roster'")
constraint_entry("child-module count vs eligible count", "3.02 vs Section 4 pages completed",
                  "structural (by construction)",
                  "Not a validation rule: the child module is nested inside each roster row "
                  "and shown if and only if that row is age-eligible, so the count of "
                  "completed child modules and the count of eligible roster rows are the same "
                  "value by construction. 3.02 itself is no longer asked -- see D-01.",
                  "The paper's separate manual count (3.02) and separate manual line-number "
                  "re-entry (4.01) were themselves the source of the mismatch risk this "
                  "requirement exists to catch",
                  "Q3 requirement 4: 'the stated number of eligible children against the "
                  "number of child modules completed'")

end_group()  # s3

# ===========================================================================
# SECTION 6: HOUSEHOLD ENVIRONMENT
# ===========================================================================
begin_group("s6", "Section 6: Household environment", "Kashi na 6: Muhallin gida",
            relevant=form_active)

q("select_one water_source", "water_source", "6.01  Main source of drinking water",
  "6.01  Babban tushen ruwan sha",
  hint_en="The source used most of the time, not an occasional backup source.",
  hint_ha="Tushen da ake amfani da shi mafi yawan lokaci, ba wanda ake amfani da shi lokaci "
          "lokaci ba.",
  required=True, appearance="minimal")
q("select_one toilet_type", "toilet_type", "6.02  Kind of toilet facility usually used",
  "6.02  Irin bayan gida da ake amfani da shi", required=True, appearance="minimal")
q("select_one yes_no", "livestock_in_compound", "6.03  Does this household keep poultry or "
  "livestock inside the compound?", "6.03  Gidan yana da kaji ko dabbobi a cikin gida?",
  required=True, appearance="quick")
q("select_one yes_no_dk", "animal_antibiotics_12m", "6.04  Have any antibiotic medicines "
  "been given to these animals in the past 12 months?", "6.04  An bawa dabbobin "
  "magungunan rigakafi cikin watanni 12 da suka wuce?",
  relevant="${livestock_in_compound}='1'", required=True, appearance="quick")
q("select_one handwash", "handwash_station", "6.05  Handwashing station with soap and "
  "water available?", "6.05  Akwai wurin wanke hannu da sabulu da ruwa?",
  hint_en="Ask to see the handwashing place -- do not just ask the question and accept a "
          "verbal answer.",
  hint_ha="A nemi ganin wurin wanke hannu -- kada kawai a tambaya a karbi amsa ta baki.",
  required=True, appearance="quick")
q("select_one yes_no_dk", "hh_diarrhoea_2w", "6.06  Has any member of this household had "
  "diarrhoea in the past two weeks?", "6.06  Wani a gida ya sha gudawa cikin makonni biyu "
  "da suka wuce?", required=True, appearance="quick")

q("select_multiple assets", "hh_assets", "6.07  Which of the following does this "
  "household own?", "6.07  Wanne daga cikin wadannan gidan yake da su?",
  hint_en="Select every item this household owns. Select 'None of these' only if none "
          "apply.",
  hint_ha="A zaba duk abin da gidan yake da shi. A zaba 'Babu ko daya' idan babu wanda ya "
          "dace.",
  required=True,
  appearance="columns-2",
  constraint="not(selected(., 'H')) or count-selected(.)=1",
  constraint_message="'None of these' cannot be selected together with any other item.")
constraint_entry("hh_assets", "6.07", "mutual exclusivity", "selected('H') implies "
                  "count-selected=1 (None of these excludes every other option)",
                  "A response set that includes both 'None of these' and a specific asset "
                  "(e.g. mobile telephone), which is self-contradictory and breaks any "
                  "asset-index construction downstream",
                  "The paper form is a 'select all that apply' multi-select with no stated "
                  "exclusivity rule for its own 'None of these' category -- identified as "
                  "defect D-09 (data that cannot be analysed as printed), resolved in form")

end_group()  # s6

# ===========================================================================
# SECTION 7: CLOSE-OUT AND SUPERVISOR REVIEW
# ===========================================================================
begin_group("s7", "Section 7: Close-out and supervisor review", "Kashi na 7: Kammalawa",
            relevant=form_active)

calc("duration_minutes", "(decimal-date-time(${end}) - decimal-date-time(${start})) * 1440")
note("interview_end_note", "Interview end time is captured automatically when you submit; "
     "the form no longer asks you to type it separately (paper 7.01). See defects report and "
     "fabrication-detection doc.", PENDING)
constraint_entry("duration_minutes", "7.01", "added field, replaces manual entry",
                  "duration_minutes = (end - start) in minutes, computed from the system "
                  "start/end metadata rather than the enumerator typing a clock time",
                  "The manually-typed 7.01 time can be edited/guessed by an enumerator "
                  "wanting to disguise a short interview; the system timestamp cannot",
                  "Directly answers Q3 requirement 11 (fabrication detection) using the "
                  "operating-conditions fraud case as the design driver")

q("text", "supervisor_note", "7.02  Observation that may help the office interpret this "
  "form", "7.02  Bayani da zai taimaka wa ofis",
  hint_en="Do not record the respondent's name or other identifying detail here -- see "
          "data-protection notes.",
  hint_ha="Kada a rubuta sunan wanda ake tambaya ko wani bayani da zai bayyana shi a nan.",
  appearance="multiline")
q("select_one enumerator", "supervisor_signoff_enum_code", "7.03  Enumerator signature "
  "(select your code again to confirm)", "7.03  Tabbatar da lambar ka", required=True,
  appearance="minimal")

end_group()  # s7

begin_group("s7sup", "Supervisor review (completed after handover, not by the enumerator)",
            "Bitar mai kula", appearance="field-list")

q("select_one enumerator", "supervisor_code", "7.04  Supervisor code", "7.04  Lambar mai "
  "kula",
  appearance="minimal")
q("select_one supervisor_decision", "supervisor_decision", "7.05  Supervisor decision on "
  "this form", "7.05  Shawarar mai kula", appearance="minimal")

end_group()

# ===========================================================================
# CHOICES
# ===========================================================================
def yn(list_name="yes_no"):
    choice(list_name, "1", "Yes", "Ee")
    choice(list_name, "2", "No", "A'a")


yn("yes_no")
choice("yes_no_dk", "1", "Yes", "Ee")
choice("yes_no_dk", "2", "No", "A'a")
choice("yes_no_dk", "8", "Do not know", "Ban sani ba")

choice("consent", "1", "Consent given", "An yarda")
choice("consent", "2", "Consent refused", "An ki")

# ---------------------------------------------------------------------------
# Below: real Hausa for every SHORT, HIGH-FREQUENCY, UNAMBIGUOUS choice list --
# family-relationship terms, water/toilet/asset categories, administrative
# decisions -- consistent with the language policy actually stated in
# documentation/06_language_and_translation.md. PENDING is reserved for full
# clinical/procedural SENTENCES (question text, constraint messages) and for
# the enumerator list, which needs no translation at all (see below), not for
# short everyday vocabulary. An earlier build broke this rule by defaulting
# every choice() call to PENDING and only overriding a handful -- caught by
# testing the form live in KoboToolbox, where a 120-option select_one
# (question 1.08) rendered the same placeholder string 120 times over in
# Hausa, the form's own default language. Fixed here, not papered over.
# ---------------------------------------------------------------------------
for code, label, ha in [("1", "Head", "Shugaban gida"), ("2", "Spouse", "Miji ko Matar aure"),
                         ("3", "Son or daughter", "Da ko 'Ya"), ("4", "Parent", "Uba ko Uwa"),
                         ("5", "Other relative", "Sauran dangi"),
                         ("6", "Not related", "Ba dangi ba")]:
    choice("relationship", code, label, ha)

for code, label, ha in [("1", "Head", "Shugaban gida"), ("2", "Spouse", "Miji ko Matar aure"),
                         ("3", "Son or daughter", "Da ko 'Ya"), ("4", "Parent", "Uba ko Uwa"),
                         ("5", "Other relative", "Sauran dangi"),
                         ("6", "Not related", "Ba dangi ba"), ("7", "Self", "Kansa/Kanta")]:
    choice("relationship_member", code, label, ha)

choice("sex", "1", "Male", "Namiji")
choice("sex", "2", "Female", "Mace")

choice("visit_result", "1", "Completed", "An kammala")
choice("visit_result", "2", "Refused", "An ki")
choice("visit_result", "3", "No competent adult after three visits",
       "Babu babba mai iya bayar da amsa bayan ziyara uku")
choice("visit_result", "4", "Dwelling vacant or demolished", "Gidan babu kowa ko an rushe shi")

choice("measure_status", "1", "Measured", "An auna")
choice("measure_status", "2", "Not measured", "Ba a auna ba")
choice("measure_status", "3", "Caregiver/child declined", "Mai kula ko yaro ya ki")

choice("measure_position", "1", "Recumbent length", "Tsawo, kwance")
choice("measure_position", "2", "Standing height", "Tsawo, a tsaye")

choice("card_seen", "1", "Card, card copy, or electronic record seen",
       "An ga katin, kwafinsa, ko bayanin lantarki")
choice("card_seen", "2", "No card seen", "Ba a ga kati ba")

choice("photo_status", "1", "Photograph taken", "An dauki hoto")
choice("photo_status", "2", "Not available", "Babu")
choice("photo_status", "3", "Caregiver declined", "Mai kula ya ki")

choice("no_specimen_reason", "1", "Caregiver refused", "Mai kula ya ki")
choice("no_specimen_reason", "2", "Child absent", "Yaro ba ya nan")
choice("no_specimen_reason", "3", "Unable to produce", "Ba a iya samu ba")
choice("no_specimen_reason", "4", "Container spoiled", "Akwatin ya lalace")
choice("no_specimen_reason", "96", "Other", "Wani")

water_opts = [
    ("Piped into dwelling", "Famfo a cikin gida"),
    ("Piped into compound", "Famfo a filin gida"),
    ("Public tap or standpipe", "Famfon jama'a"),
    ("Tube well or borehole", "Rijiyar bututu"),
    ("Protected dug well", "Rijiya mai kariya"),
    ("Unprotected dug well", "Rijiya marar kariya"),
    ("Protected spring", "Maɓulɓula mai kariya"),
    ("Unprotected spring", "Maɓulɓula marar kariya"),
    ("Rainwater", "Ruwan sama"),
    ("Tanker or cart", "Ruwan tanka ko keken ruwa"),
    ("Surface water", "Ruwan bude (kogi ko tafki)"),
]
for i, (label, ha) in enumerate(water_opts, start=1):
    choice("water_source", str(i), label, ha)

toilet_opts = [
    ("Flush to sewer", "Fasa ruwa zuwa magudanar ruwa"),
    ("Flush to septic tank", "Fasa ruwa zuwa tankin najasa"),
    ("Flush to pit latrine", "Fasa ruwa zuwa rami"),
    ("Ventilated improved pit", "Rami mai iska (VIP)"),
    ("Pit latrine with slab", "Rami mai bene"),
    ("Pit latrine without slab", "Rami marar bene"),
    ("Composting toilet", "Bayan gida na taki"),
    ("Bucket", "Bokiti"),
    ("No facility or bush", "Babu wurin, daji"),
]
for i, (label, ha) in enumerate(toilet_opts, start=1):
    choice("toilet_type", str(i), label, ha)

choice("handwash", "1", "Observed, soap and water", "An gani, akwai sabulu da ruwa")
choice("handwash", "2", "Reported only, not observed", "An fada kawai, ba a gani ba")
choice("handwash", "3", "Not present", "Babu")

for code, label, ha in [("A", "Radio", "Rediyo"), ("B", "Television", "Talabijin"),
                         ("C", "Mobile telephone", "Wayar hannu"), ("D", "Bicycle", "Keke"),
                         ("E", "Motorcycle", "Babur"), ("F", "Car or truck", "Mota ko babbar mota"),
                         ("G", "Refrigerator", "Firji"),
                         ("H", "None of these", "Babu ko daya daga cikin wadannan")]:
    choice("assets", code, label, ha)

choice("supervisor_decision", "1", "Accept", "An amince")
choice("supervisor_decision", "2", "Return for correction", "A mayar don gyara")
choice("supervisor_decision", "3", "Void", "An soke")

# Enumerator choice list sourced from staff_roster.csv (verified 120 rows,
# 24 teams x 5 staff, team codes 1:1 with specimen_label_allocation.csv).
# Small enough (120 rows) to sit in the survey's own choices sheet safely --
# unlike settlements (2,524 rows), this does not need external-file treatment.
# The Hausa column deliberately repeats the English label rather than PENDING:
# "Enumerator 003 (TM03, Ilela)" is a code/team identifier, not prose -- it
# needs no translation, and defaulting it to a placeholder was the actual bug
# (see the block comment above).
import csv as _csv
_staff_path = os.path.join(ROOT, "form", "media", "staff_roster.csv")
with open(_staff_path, encoding="utf-8-sig") as f:
    for r in _csv.DictReader(f):
        _enum_label = f"{r['label']} ({r['team_code']}, {r['assigned_lga']})"
        choice("enumerator", r["name"], _enum_label, _enum_label)

# No medicine_list choices sheet. Question 4.13's coded medicine list is
# missing from the data pack (defect D-05) and is not something to invent a
# substitute for, sourced or not -- antibiotic_code (above) is a plain text
# field instead, and the gap is reported in documentation/01_defects_report.md
# and constraint_register.csv, not filled in.

print(f"{len(rows)} survey rows, {len(choices)} choice rows, {len(register)} register rows")

# ===========================================================================
# WRITE THE XLSFORM WORKBOOK
# ===========================================================================
SURVEY_COLS = [
    "type", "name", "label::English (en)", "label::Hausa (ha)",
    "hint::English (en)", "hint::Hausa (ha)",
    "guidance_hint::English (en)", "guidance_hint::Hausa (ha)",
    "required", "relevant",
    "constraint", "constraint_message::English (en)", "calculation",
    "appearance", "default", "read_only", "choice_filter",
]
CHOICE_COLS = ["list_name", "name", "label::English (en)", "label::Hausa (ha)"]
SETTINGS_COLS = list(settings.keys())

wb = openpyxl.Workbook()
ws_survey = wb.active
ws_survey.title = "survey"
ws_survey.append(SURVEY_COLS)
for r in rows:
    ws_survey.append([r.get(c, "") for c in SURVEY_COLS])

ws_choices = wb.create_sheet("choices")
ws_choices.append(CHOICE_COLS)
for r in choices:
    ws_choices.append([r.get(c, "") for c in CHOICE_COLS])

ws_settings = wb.create_sheet("settings")
ws_settings.append(SETTINGS_COLS)
ws_settings.append([settings[c] for c in SETTINGS_COLS])

os.makedirs(os.path.dirname(FORM_PATH), exist_ok=True)
wb.save(FORM_PATH)
print(f"wrote {FORM_PATH}")

# ===========================================================================
# WRITE THE CONSTRAINT REGISTER
# ===========================================================================
REG_COLS = ["id", "field_name", "questionnaire_no", "rule_type", "rule_added",
            "prevents", "source_or_judgement"]
with open(REGISTER_PATH, "w", newline="", encoding="utf-8") as f:
    w = csv.DictWriter(f, fieldnames=REG_COLS)
    w.writeheader()
    for r in register:
        w.writerow(r)
print(f"wrote {REGISTER_PATH} ({len(register)} rows)")

